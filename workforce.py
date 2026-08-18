"""
Camada de dados de Recursos Humanos / Efetivo do Dashboard Qualitech.

Le o arquivo "Planejamento_Ativos_Status_*.xlsx" (abas: Ativos, Turnover,
Novos, Bloqueios, END) e devolve um WorkforceData com os indicadores de
pessoas prontos para os cards/gráficos do app.py — em paralelo aos
indicadores operacionais (diárias/WO) que já vêm de data.py.

Segue o mesmo estilo de leitura tolerante de data.py: lê até encontrar uma
linha vazia / "Total", em vez de fixar o número de linhas, para aguentar
pequenas variações de mês a mês.
"""
from __future__ import annotations

import datetime as dt
import glob
import os
from dataclasses import dataclass, field

import openpyxl
import pandas as pd

# --------------------------------------------------------------------------- #
# Descoberta de arquivo
# --------------------------------------------------------------------------- #

_KEYWORDS = ("ativos", "planejamento")


def find_workforce_workbook(base_dir: str) -> str | None:
    """Localiza o Planejamento_Ativos_*.xlsx mais recente dentro de base_dir
    (mesma pasta sincronizada usada para o Performance_*.xlsx)."""
    candidates = list_available_workforce_workbooks(base_dir)
    return candidates[0] if candidates else None


def list_available_workforce_workbooks(base_dir: str) -> list[str]:
    if not base_dir or not os.path.isdir(base_dir):
        return []
    all_xlsx = glob.glob(os.path.join(base_dir, "**", "*.xlsx"), recursive=True)
    all_xlsx = [c for c in all_xlsx if not os.path.basename(c).startswith("~$")]
    candidates = [c for c in all_xlsx if any(k in os.path.basename(c).lower() for k in _KEYWORDS)]
    candidates.sort(key=os.path.getmtime, reverse=True)
    return candidates


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def _norm_coord(v):
    if not isinstance(v, str):
        return None
    v = v.strip()
    if not v:
        return None
    fix = {"vitoria": "Vitória", "marcelo": "Marcelo"}
    key = v.lower()
    if key in fix:
        return fix[key]
    # exclude non-coordinator "situação" values that leak into the column
    if key in ("desligado", "inss", "tbd"):
        return None
    return v


def _norm_tipo_mo(v):
    if not isinstance(v, str):
        return "Outros"
    v = v.lower()
    has_insp = "inspe" in v
    has_manu = "manuten" in v
    if has_insp and not has_manu:
        return "Inspeção"
    if has_manu and not has_insp:
        return "Manutenção"
    return "Outros"


def _norm_contrato(v):
    if not isinstance(v, str):
        return None
    v = v.strip().upper()
    return {"FIXA": "Fixa", "VARIÁVEL": "Variável", "VARIAVEL": "Variável", "SPOT": "Spot"}.get(v)


def _parse_date(v):
    if isinstance(v, dt.datetime):
        return v
    if isinstance(v, str):
        v = v.strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%y"):
            try:
                return dt.datetime.strptime(v, fmt)
            except ValueError:
                continue
    return None


MESES_PT = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}


# --------------------------------------------------------------------------- #
# Estrutura de dados devolvida
# --------------------------------------------------------------------------- #

@dataclass
class WorkforceData:
    path: str
    mtime: float
    periodo_label: str

    ativos: pd.DataFrame = field(default_factory=pd.DataFrame)
    por_tipo_mo: pd.DataFrame = field(default_factory=pd.DataFrame)
    por_contrato: pd.DataFrame = field(default_factory=pd.DataFrame)
    por_coordenador: pd.DataFrame = field(default_factory=pd.DataFrame)
    por_cliente: pd.DataFrame = field(default_factory=pd.DataFrame)
    turnover: pd.DataFrame = field(default_factory=pd.DataFrame)
    novos: pd.DataFrame = field(default_factory=pd.DataFrame)
    certs: pd.DataFrame = field(default_factory=pd.DataFrame)

    headcount_total: int = 0
    headcount_ativo: int = 0
    headcount_inativo: int = 0
    headcount_afastado: int = 0
    lideres_n: int = 0
    bloqueios_n: int = 0

    @property
    def pct_lideres(self) -> float:
        return (self.lideres_n / self.headcount_ativo) if self.headcount_ativo else 0.0

    @property
    def novos_admitidos_n(self) -> int:
        return int(self.novos["nome"].count()) if not self.novos.empty else 0

    @property
    def turnover_last(self) -> dict | None:
        if self.turnover.empty:
            return None
        return self.turnover.iloc[-1].to_dict()

    @property
    def turnover_rate_last(self) -> float | None:
        last = self.turnover_last
        if not last or not last.get("headcount_inicial"):
            return None
        saidas = (last.get("demitidos") or 0) + (last.get("pedido_demissao") or 0)
        return saidas / last["headcount_inicial"]

    @property
    def certs_vencendo_90d_n(self) -> int:
        if self.certs.empty:
            return 0
        today = dt.datetime.now()
        horizon = today + dt.timedelta(days=90)
        v = self.certs["validade_dt"].dropna()
        return int(((v >= today) & (v <= horizon)).sum())

    @property
    def certs_vencidas_n(self) -> int:
        if self.certs.empty:
            return 0
        today = dt.datetime.now()
        v = self.certs["validade_dt"].dropna()
        return int((v < today).sum())

    @property
    def certs_criticas(self) -> pd.DataFrame:
        """Certificações já vencidas ou vencendo nos próximos 90 dias, ordenadas
        por urgência — para a tabela de risco de compliance."""
        if self.certs.empty:
            return pd.DataFrame()
        today = dt.datetime.now()
        horizon = today + dt.timedelta(days=90)
        c = self.certs.dropna(subset=["validade_dt"])
        c = c[c["validade_dt"] <= horizon].copy()
        c["dias_restantes"] = (c["validade_dt"] - today).dt.days
        return c.sort_values("dias_restantes")


# --------------------------------------------------------------------------- #
# Loader principal
# --------------------------------------------------------------------------- #

def load_workforce_data(path: str) -> WorkforceData:
    wb = openpyxl.load_workbook(path, data_only=True)
    mtime = os.path.getmtime(path)
    periodo_label = os.path.splitext(os.path.basename(path))[0]

    # ---- Ativos: roster principal ----
    ativos_df = pd.DataFrame()
    headcount_total = headcount_ativo = headcount_inativo = headcount_afastado = 0
    lideres_n = 0
    por_tipo_mo_df = pd.DataFrame(columns=["tipo", "n"])
    por_contrato_df = pd.DataFrame(columns=["tipo", "n"])
    por_coord_df = pd.DataFrame(columns=["coordenador", "n"])
    por_cliente_df = pd.DataFrame(columns=["cliente", "n"])

    if "Ativos" in wb.sheetnames:
        ws = wb["Ativos"]
        cols = ["matricula", "nome", "funcao", "funcao_abrev", "tipo_mo", "situacao",
                "coordenador", "sispat", "cliente", "wo", "tipo_contrato", "plataforma",
                "residencia", "lider", "admissao", "lideranca_projeto"]
        rows = []
        empty_streak = 0
        r = 2
        while r < 2000 and empty_streak < 20:
            matricula = ws.cell(r, 2).value
            if matricula in (None, ""):
                empty_streak += 1
                r += 1
                continue
            empty_streak = 0
            rows.append([ws.cell(r, c).value for c in range(2, 18)])
            r += 1
        ativos_df = pd.DataFrame(rows, columns=cols)

        headcount_total = len(ativos_df)
        situ = ativos_df["situacao"].astype(str).str.strip()
        headcount_ativo = int((situ == "Ativo").sum())
        headcount_inativo = int((situ == "Inativo").sum())
        headcount_afastado = int((situ == "Afastado").sum())

        ativo_df = ativos_df[situ == "Ativo"].copy()
        lideres_n = int((ativo_df["lider"] == True).sum())  # noqa: E712

        tipo_mo_counts = ativo_df["tipo_mo"].apply(_norm_tipo_mo).value_counts()
        order_mo = [t for t in ["Inspeção", "Manutenção", "Outros"] if t in tipo_mo_counts.index]
        por_tipo_mo_df = pd.DataFrame({"tipo": order_mo, "n": [int(tipo_mo_counts[t]) for t in order_mo]})

        contrato_counts = ativo_df["tipo_contrato"].apply(_norm_contrato).dropna().value_counts()
        order_c = [t for t in ["Fixa", "Variável", "Spot"] if t in contrato_counts.index]
        por_contrato_df = pd.DataFrame({"tipo": order_c, "n": [int(contrato_counts[t]) for t in order_c]})

        coord_counts = ativo_df["coordenador"].apply(_norm_coord).dropna().value_counts().sort_values(ascending=False)
        por_coord_df = pd.DataFrame({"coordenador": coord_counts.index, "n": coord_counts.values})

        cli_counts = ativo_df["cliente"].dropna().value_counts().sort_values(ascending=False)
        top = cli_counts.head(8)
        outros = int(cli_counts.iloc[8:].sum())
        cli_labels = list(top.index) + (["Outros"] if outros else [])
        cli_vals = list(top.values) + ([outros] if outros else [])
        por_cliente_df = pd.DataFrame({"cliente": cli_labels, "n": cli_vals})

    # ---- Turnover: evolução mensal de headcount ----
    turnover_df = pd.DataFrame()
    if "Turnover" in wb.sheetnames:
        ws = wb["Turnover"]
        header_row = [ws.cell(2, c).value for c in range(2, ws.max_column + 1)]
        labels, col_idx = [], []
        for i, h in enumerate(header_row):
            c = i + 2
            if not isinstance(h, str):
                continue
            h_low = h.lower().strip()
            if "total" in h_low:
                continue
            labels.append(h)
            col_idx.append(c)

        field_rows = {
            "headcount_inicial": 3, "admitidos": 4, "demitidos": 5,
            "transferencia": 6, "pedido_demissao": 7, "saldo_final": 8,
        }
        data = {"label": labels}
        for field_name, row_n in field_rows.items():
            data[field_name] = [ws.cell(row_n, c).value or 0 for c in col_idx]
        turnover_df = pd.DataFrame(data)

    # ---- Novos: pipeline de admissões recentes ----
    novos_df = pd.DataFrame()
    if "Novos" in wb.sheetnames:
        ws = wb["Novos"]
        rows = []
        for r in range(4, ws.max_row + 1):
            admissao, nome, cargo, coord, liberacao = [ws.cell(r, c).value for c in (6, 7, 8, 9, 10)]
            if not nome:
                continue
            rows.append(dict(admissao=admissao, nome=nome, cargo=cargo,
                              coordenador=_norm_coord(coord) or coord, liberacao=liberacao))
        novos_df = pd.DataFrame(rows)

    # ---- Bloqueios ----
    bloqueios_n = 0
    if "Bloqueios" in wb.sheetnames:
        ws = wb["Bloqueios"]
        for r in range(4, ws.max_row + 1):
            if ws.cell(r, 3).value:
                bloqueios_n += 1

    # ---- END: certificações / qualificações ----
    certs_df = pd.DataFrame()
    if "END" in wb.sheetnames:
        ws = wb["END"]
        rows = []
        for r in range(2, ws.max_row + 1):
            nome, funcao, admissao, situacao, qualif, emissao, validade = [ws.cell(r, c).value for c in range(1, 8)]
            if not nome:
                continue
            rows.append(dict(
                nome=str(nome).strip(), funcao=funcao, situacao=situacao,
                qualificacao=str(qualif).strip() if qualif else None,
                emissao_dt=_parse_date(emissao), validade_dt=_parse_date(validade),
            ))
        certs_df = pd.DataFrame(rows)

    return WorkforceData(
        path=path, mtime=mtime, periodo_label=periodo_label,
        ativos=ativos_df, por_tipo_mo=por_tipo_mo_df, por_contrato=por_contrato_df,
        por_coordenador=por_coord_df, por_cliente=por_cliente_df,
        turnover=turnover_df, novos=novos_df, certs=certs_df,
        headcount_total=headcount_total, headcount_ativo=headcount_ativo,
        headcount_inativo=headcount_inativo, headcount_afastado=headcount_afastado,
        lideres_n=lideres_n, bloqueios_n=bloqueios_n,
    )
