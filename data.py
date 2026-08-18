"""
Camada de dados do Dashboard Qualitech.

Le o arquivo Performance_<Mes>.xlsx (o mesmo layout de abas usado hoje:
Pivot, Tabelas, Taxa_Utilização, SISPAT, Evolução_DIARIO/ACUMULADO,
Planilha1, Plan_Estretégico, Historico) e devolve um dicionário de
DataFrames/valores já tratados, prontos para os gráficos do app.py.

Escrito para ser resiliente a pequenas variações de layout mes a mes:
- descobre onde a tabela termina lendo até encontrar a linha "Total Geral"
  (ou a primeira linha totalmente vazia), em vez de fixar o numero de linhas.
- todas as fórmulas/valores são lidos com data_only=True (valores calculados
  pelo Excel), então o arquivo precisa ter sido salvo ao menos uma vez no
  Excel/OneDrive antes de ser lido aqui (o que já é o caso de um arquivo
  sincronizado normalmente).
"""
from __future__ import annotations

import datetime as dt
import glob
import os
from dataclasses import dataclass, field

import openpyxl
import pandas as pd


# --------------------------------------------------------------------------- #
# Descoberta de arquivo
# --------------------------------------------------------------------------- #

def find_latest_workbook(base_dir: str) -> str | None:
    """
    Localiza o Performance_*.xlsx mais recente dentro de base_dir.

    Aceita tanto:
      base_dir apontando direto para a pasta do mês (ex.: .../9. Performance/Ago-2026)
      quanto
      base_dir apontando para a pasta pai (ex.: .../9. Performance), caso em
      que procura em todas as subpastas e pega o arquivo modificado mais
      recentemente.
    """
    if not base_dir or not os.path.isdir(base_dir):
        return None

    candidates = glob.glob(os.path.join(base_dir, "**", "*.xlsx"), recursive=True)
    candidates = [c for c in candidates if not os.path.basename(c).startswith("~$")]
    if not candidates:
        return None
    candidates.sort(key=os.path.getmtime, reverse=True)
    return candidates[0]


def list_available_workbooks(base_dir: str) -> list[str]:
    if not base_dir or not os.path.isdir(base_dir):
        return []
    candidates = glob.glob(os.path.join(base_dir, "**", "*.xlsx"), recursive=True)
    candidates = [c for c in candidates if not os.path.basename(c).startswith("~$")]
    candidates.sort(key=os.path.getmtime, reverse=True)
    return candidates


# --------------------------------------------------------------------------- #
# Helpers de leitura tolerante a tamanho de tabela
# --------------------------------------------------------------------------- #

def _rows(ws, min_row, max_row, min_col, max_col=None):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        yield [c.value for c in row]


def _read_until_total(ws, start_row, min_col, max_col, label_col_offset=0, max_scan=600):
    """Lê linhas a partir de start_row até achar 'Total Geral' na 1a coluna
    (ou uma linha totalmente vazia), sem depender de um número fixo de linhas."""
    out = []
    for r in range(start_row, start_row + max_scan):
        row = [ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
        label = row[label_col_offset]
        if label is None and all(v is None for v in row):
            break
        if isinstance(label, str) and "total" in label.lower():
            break
        out.append(row)
    return out


# --------------------------------------------------------------------------- #
# Estrutura de dados devolvida
# --------------------------------------------------------------------------- #

@dataclass
class DashboardData:
    path: str
    mtime: float
    periodo_label: str

    wo: pd.DataFrame = field(default_factory=pd.DataFrame)
    por_coordenador: pd.DataFrame = field(default_factory=pd.DataFrame)
    por_tipo: pd.DataFrame = field(default_factory=pd.DataFrame)
    por_cliente: pd.DataFrame = field(default_factory=pd.DataFrame)
    backlog_coordenador: pd.DataFrame = field(default_factory=pd.DataFrame)
    sispat: pd.DataFrame = field(default_factory=pd.DataFrame)
    utilizacao_diaria: pd.DataFrame = field(default_factory=pd.DataFrame)
    utilizacao_mensal: pd.DataFrame = field(default_factory=pd.DataFrame)
    evolucao_diaria: pd.DataFrame = field(default_factory=pd.DataFrame)
    evolucao_acumulada: pd.DataFrame = field(default_factory=pd.DataFrame)
    plano_estrategico: pd.DataFrame = field(default_factory=pd.DataFrame)
    movimentacao_clientes: pd.DataFrame = field(default_factory=pd.DataFrame)

    lancado_total: float = 0.0
    pendente_total: float = 0.0

    @property
    def totais(self) -> dict:
        if self.por_coordenador.empty:
            return dict(planejado=0, programado=0, real_prog=0, real_acum=0)
        c = self.por_coordenador
        return dict(
            planejado=c["planejado"].sum(),
            programado=c["programado"].sum(),
            real_prog=c["real_prog"].sum(),
            real_acum=c["real_acum"].sum(),
        )

    @property
    def n_wo(self) -> int:
        return int(self.wo["wo"].count()) if not self.wo.empty else 0

    @property
    def por_tipo_wo(self) -> pd.DataFrame:
        if self.wo.empty:
            return pd.DataFrame(columns=["tipo_wo", "n", "pct"])
        counts = self.wo["tipo_wo"].value_counts()
        total = counts.sum()
        order = [t for t in ["Fixo", "Variável", "Spot"] if t in counts.index] + [
            t for t in counts.index if t not in ("Fixo", "Variável", "Spot")
        ]
        df = pd.DataFrame({"tipo_wo": order, "n": [int(counts[t]) for t in order]})
        df["pct"] = df["n"] / total if total else 0
        return df

    @property
    def cross_tipowo_tipo(self) -> pd.DataFrame:
        if self.wo.empty:
            return pd.DataFrame()
        main_tipos = ["Inspeção", "Manutenção"]
        pivot = pd.crosstab(self.wo["tipo_wo"], self.wo["tipo"])
        for t in main_tipos:
            if t not in pivot.columns:
                pivot[t] = 0
        pivot["Outros"] = pivot.drop(columns=main_tipos, errors="ignore").sum(axis=1) - pivot.get("Outros", 0)
        pivot = pivot[main_tipos + ["Outros"]]
        order = [t for t in ["Fixo", "Variável", "Spot"] if t in pivot.index]
        return pivot.loc[order]

    @property
    def sispat_summary(self) -> dict:
        if self.sispat.empty:
            return dict(sim=0, nao=0, total=0)
        sim = int((self.sispat["sispat"] == "Sim").sum())
        nao = int((self.sispat["sispat"] == "Não").sum())
        return dict(sim=sim, nao=nao, total=sim + nao)

    @property
    def pct_utilizacao_ociosidade(self) -> tuple[float | None, float | None, dict | None]:
        if self.utilizacao_diaria.empty:
            return None, None, None
        valid = self.utilizacao_diaria.dropna(subset=["taxa"])
        if valid.empty:
            return None, None, None
        last = valid.iloc[-1]
        return float(last["taxa"]), 1 - float(last["taxa"]), last.to_dict()

    @property
    def ildd_medio(self) -> float | None:
        if self.wo.empty or "ildd" not in self.wo:
            return None
        v = self.wo["ildd"].dropna()
        return float(v.mean()) if len(v) else None

    @property
    def iap_medio(self) -> float | None:
        if self.wo.empty or "iap" not in self.wo:
            return None
        v = self.wo["iap"].dropna()
        return float(v.mean()) if len(v) else None


# --------------------------------------------------------------------------- #
# Loader principal
# --------------------------------------------------------------------------- #

def load_workbook_data(path: str) -> DashboardData:
    wb = openpyxl.load_workbook(path, data_only=True)
    mtime = os.path.getmtime(path)

    # ---- Pivot: WO detail ----
    wo_rows = []
    if "Pivot" in wb.sheetnames:
        ws = wb["Pivot"]
        for r in _rows(ws, 4, 600, 2, 16):
            coord, cliente, plataforma, wo_id, tipo_wo, tipo, planejado, programado, ajuste_mos, \
                real_prog, real_acum, lancado, ajuste_mes, ildd, iap = r
            if coord is None:
                continue
            if isinstance(coord, str) and "total" in coord.lower():
                break
            wo_rows.append(dict(
                coord=coord, cliente=cliente, plataforma=plataforma, wo=wo_id,
                tipo_wo=tipo_wo, tipo=tipo,
                planejado=planejado or 0, programado=programado or 0,
                real_prog=real_prog or 0, real_acum=real_acum or 0,
                lancado=lancado or 0, ajuste_mes=ajuste_mes or 0,
                ildd=ildd, iap=iap,
            ))
    wo_df = pd.DataFrame(wo_rows)

    # ---- Tabelas: coordenador / tipo / cliente ----
    por_coord_df = pd.DataFrame()
    por_tipo_df = pd.DataFrame()
    por_cliente_df = pd.DataFrame()
    lancado_total = pendente_total = 0.0
    if "Tabelas" in wb.sheetnames:
        ws = wb["Tabelas"]
        coord_rows = _read_until_total(ws, 3, 2, 8, label_col_offset=0)
        por_coord_df = pd.DataFrame(coord_rows, columns=[
            "coord", "planejado", "programado", "real_prog", "real_acum", "lancado", "ajuste_mes"
        ]).dropna(subset=["coord"])

        tipo_rows = _read_until_total(ws, 3, 12, 16, label_col_offset=0)
        por_tipo_df = pd.DataFrame(tipo_rows, columns=[
            "tipo", "planejado", "programado", "real_prog", "real_acum"
        ]).dropna(subset=["tipo"])

        cli_rows = _read_until_total(ws, 17, 2, 9, label_col_offset=0)
        por_cliente_df = pd.DataFrame(cli_rows, columns=[
            "cliente", "planejado", "programado", "real_prog", "real_acum", "lancado", "ajuste_mes", "pct_diarias"
        ]).dropna(subset=["cliente"])

        lancado_total = ws["F13"].value or 0
        pendente_total = ws["F14"].value or 0

    # ---- Planilha1: backlog logística ----
    backlog_df = pd.DataFrame()
    if "Planilha1" in wb.sheetnames:
        ws = wb["Planilha1"]
        rows = _read_until_total(ws, 3, 2, 4, label_col_offset=0)
        backlog_df = pd.DataFrame(rows, columns=["coord", "programado", "backlog_logistica"]).dropna(subset=["coord"])

    # ---- SISPAT ----
    sispat_df = pd.DataFrame()
    if "SISPAT" in wb.sheetnames:
        ws = wb["SISPAT"]
        recs, cur_cliente = [], None
        for r in _rows(ws, 2, 200, 3, 5):
            cliente, plataforma, sim = r
            if cliente:
                cur_cliente = cliente
            if plataforma is None:
                continue
            recs.append(dict(cliente=cur_cliente, plataforma=plataforma, sispat=sim))
        sispat_df = pd.DataFrame(recs)

    # ---- Taxa_Utilização: diária + mensal ----
    util_diaria_df = pd.DataFrame()
    util_mensal_df = pd.DataFrame()
    if "Taxa_Utilização" in wb.sheetnames:
        ws = wb["Taxa_Utilização"]
        recs = []
        for r in _rows(ws, 4, 40, 2, 8):
            date, real_prog, real_acum, ativo, b2b, taxa, drake = r
            if not isinstance(date, dt.datetime):
                continue
            recs.append(dict(date=date, real_prog=real_prog, real_acum=real_acum, ativo=ativo, b2b=b2b, taxa=taxa))
        util_diaria_df = pd.DataFrame(recs)

        recs2 = []
        for r in _rows(ws, 40, 80, 2, 9):
            mes, prog, real, ativo, b2b, dias, cap, taxa = r
            if not isinstance(mes, dt.datetime) or real is None:
                continue
            recs2.append(dict(mes=mes, programado=prog, realizado=real, ativo=ativo, b2b=b2b,
                               dias=dias, capacidade=cap, taxa=taxa))
        util_mensal_df = pd.DataFrame(recs2)

    # ---- Evolução diária / acumulada ----
    evo_d_df = pd.DataFrame()
    if "Evolução_DIARIO" in wb.sheetnames:
        ws = wb["Evolução_DIARIO"]
        recs = []
        for r in _rows(ws, 3, 40, 2, 7):
            date, prog, real_prog, logistica, real_acum, dif = r
            if not isinstance(date, dt.datetime):
                continue
            recs.append(dict(date=date, dia=date.day, programado=prog, real_prog=real_prog,
                              logistica=logistica, real_acum=real_acum, dif=dif))
        evo_d_df = pd.DataFrame(recs)

    evo_a_df = pd.DataFrame()
    if "Evolução_ACUMULADO" in wb.sheetnames:
        ws = wb["Evolução_ACUMULADO"]
        recs = []
        for r in _rows(ws, 3, 40, 2, 8):
            date, prog_ac, real_prog_ac, logistica, real_ac, *_ = r
            if not isinstance(date, dt.datetime):
                continue
            recs.append(dict(date=date, dia=date.day, programado_acum=prog_ac,
                              real_prog_acum=real_prog_ac, logistica_acum=logistica, real_acum=real_ac))
        evo_a_df = pd.DataFrame(recs)

    # ---- Plano Estratégico ----
    plano_df = pd.DataFrame()
    if "Plan_Estretégico" in wb.sheetnames:
        ws = wb["Plan_Estretégico"]
        recs = []
        for r in _rows(ws, 4, 40, 7, 14):
            mes, meta, comercial, efetivo, capacidade, buffer, real, dd = r
            if mes is None:
                continue
            recs.append(dict(mes=mes, meta=meta, comercial=comercial, efetivo_ativo=efetivo,
                              capacidade=capacidade, buffer=buffer, real=real, dd=dd))
        plano_df = pd.DataFrame(recs)

    # ---- Historico: movimentação de clientes ----
    hist_df = pd.DataFrame()
    if "Historico" in wb.sheetnames:
        ws = wb["Historico"]
        recs = []
        for r in _rows(ws, 4, 40, 27, 29):
            cliente, entraram, sairam = r
            if cliente is None or (isinstance(cliente, str) and cliente.lower() in ("redução", "reducao")):
                continue
            recs.append(dict(cliente=cliente, entraram=entraram or 0, sairam=sairam or 0))
        hist_df = pd.DataFrame(recs)

    periodo_label = os.path.splitext(os.path.basename(path))[0]

    return DashboardData(
        path=path, mtime=mtime, periodo_label=periodo_label,
        wo=wo_df, por_coordenador=por_coord_df, por_tipo=por_tipo_df, por_cliente=por_cliente_df,
        backlog_coordenador=backlog_df, sispat=sispat_df,
        utilizacao_diaria=util_diaria_df, utilizacao_mensal=util_mensal_df,
        evolucao_diaria=evo_d_df, evolucao_acumulada=evo_a_df,
        plano_estrategico=plano_df, movimentacao_clientes=hist_df,
        lancado_total=lancado_total, pendente_total=pendente_total,
    )
